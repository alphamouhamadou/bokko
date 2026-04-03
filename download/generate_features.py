import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Fonctionnalités BOKKO"

# Colors
green_primary = "006233"
green_light = "E8F5E9"
green_medium = "C8E6C9"
yellow_accent = "FFD700"
white = "FFFFFF"
grey_dark = "333333"
grey_light = "F5F5F5"
grey_alt = "E9E9E9"
red_priority = "FFD700"
orange_priority = "FF9800"

# Styles
title_font = Font(name='Times New Roman', size=20, bold=True, color=green_primary)
subtitle_font = Font(name='Times New Roman', size=12, italic=True, color="666666")
header_font = Font(name='Times New Roman', size=11, bold=True, color=white)
header_fill = PatternFill(start_color=green_primary, end_color=green_primary, fill_type="solid")
category_font = Font(name='Times New Roman', size=12, bold=True, color=green_primary)
category_fill = PatternFill(start_color=green_light, end_color=green_light, fill_type="solid")
normal_font = Font(name='Times New Roman', size=11, color="333333")
desc_font = Font(name='Times New Roman', size=10, color="555555")
p1_font = Font(name='Times New Roman', size=10, bold=True, color="D32F2F")
p2_font = Font(name='Times New Roman', size=10, bold=True, color="E65100")
p3_font = Font(name='Times New Roman', size=10, bold=True, color="1565C0")
p4_font = Font(name='Times New Roman', size=10, bold=True, color="6A1B9A")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)

# Column widths
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 6
ws.column_dimensions['C'].width = 36
ws.column_dimensions['D'].width = 58
ws.column_dimensions['E'].width = 52
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14

# Title
ws.merge_cells('B2:G2')
ws['B2'] = "BOKKO - Fonctionnalités à Ajouter"
ws['B2'].font = title_font
ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 38

ws.merge_cells('B3:G3')
ws['B3'] = "Roadmap de développement - Covoiturage & Colis au Sénégal"
ws['B3'].font = subtitle_font
ws['B3'].alignment = Alignment(horizontal='left', vertical='center')

# Legend
row = 5
ws.merge_cells(f'B{row}:G{row}')
ws[f'B{row}'] = "LÉGENDE DES PRIORITÉS"
ws[f'B{row}'].font = Font(name='Times New Roman', size=10, bold=True, color=grey_dark)
row = 6
for label, font in [("P1 - Critique", p1_font), ("P2 - Important", p2_font), ("P3 - Moyen", p3_font), ("P4 - Futur", p4_font)]:
    ws[f'B{row}'] = label
    ws[f'B{row}'].font = font
    ws[f'B{row}'].alignment = left_align
    row += 1

# Headers
row = 9
headers = ["#", "Fonctionnalité", "Description", "Détails techniques", "Priorité", "Complexité"]
for i, h in enumerate(headers):
    col = i + 2
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[row].height = 28

# Features data organized by category
features = [
    # Category: SECURITÉ & AUTH
    ("CATEGORIE", "🔐 SÉCURITÉ & AUTHENTIFICATION", None),
    ("F", "Mot de passe oublié", "Permettre à l'utilisateur de réinitialiser son mot de passe via SMS ou code OTP envoyé sur son numéro de téléphone.", "API /api/auth/reset-password + API /api/auth/verify-otp + modification du schéma User (add resetToken, resetExpires) + composant ForgotPasswordForm", "P1", "Moyenne"),
    ("F", "Modification du mot de passe", "L'utilisateur connecté peut changer son mot de passe actuel depuis un écran de paramètres de compte.", "API /api/auth/change-password + composant ChangePasswordForm dans settings + validation ancien/nouveau mot de passe", "P1", "Faible"),
    ("F", "Suppression de compte", "Permettre à l'utilisateur de supprimer définitivement son compte et toutes ses données associées (réservations, colis, avis).", "API /api/auth/delete-account + soft-delete User (add deletedAt) + composant DeleteAccountButton + confirmation modal", "P2", "Moyenne"),
    ("F", "Limitation tentatives de connexion", "Bloquer le compte temporairement après 5 tentatives de connexion échouées pour prévenir les attaques brute-force.", "Middleware API + champs failedAttempts, lockedUntil dans User + logique de vérification dans /api/auth/login", "P1", "Faible"),
    ("F", "Durée de session / Auto-déconnexion", "Déconnecter automatiquement l'utilisateur après une période d'inactivité configurable (ex: 7 jours) pour la sécurité.", "Champs lastActivity dans User + middleware de vérification dans chaque API + hook useEffect d'auto-check côté client", "P2", "Faible"),

    # Category: GÉOLOCALISATION & CARTE
    ("CATEGORIE", "🗺️ GÉOLOCALISATION & CARTE", None),
    ("F", "Suivi GPS en temps réel", "Le chauffeur partage sa position en temps réel pendant le trajet. Les passagers voient la progression sur une carte.", "API /api/trips/[id]/location (PATCH) + WebSocket ou polling 10s + composant carte Leaflet/Mapbox + permissions géolocalisation", "P1", "Élevée"),
    ("F", "Points de rassemblement sur carte", "Afficher une carte interactive avec les points de départ/arrivée et les points de rassemblement courants pour chaque ville.", "Intégration Leaflet (gratuit) + markers pour Thiès/Dakar/Thiènaba + sous-locations géocodées + composant MapPicker", "P2", "Moyenne"),
    ("F", "Itinéraire visuel sur carte", "Afficher le tracé routier entre l'origine et la destination sur la carte avec les étapes principales.", "Leaflet + API OSRM (gratuit) pour le tracé + polyline sur carte + calcul distance/durée estimée", "P2", "Moyenne"),
    ("F", "Géolocalisation du passager", "Détecter automatiquement la position du passager pour proposer la ville de départ la plus proche.", "navigator.geolocation + reverse geocoding Nominatim (gratuit) + auto-sélection dans le formulaire de recherche", "P3", "Faible"),

    # Category: NOTIFICATIONS
    ("CATEGORIE", "🔔 NOTIFICATIONS PUSH", None),
    ("F", "Notifications push web", "Envoyer des notifications push même quand l'app est fermée (nouvelle réservation, confirmation, rappel de départ, etc.).", "Service Worker + Push API + API /api/push/subscribe + Web Push Protocol + VAPID keys + cron pour les rappels", "P1", "Élevée"),
    ("F", "Rappel avant le départ", "Notifier le passager 30 min avant l'heure de départ de son trajet réservé pour éviter les retards.", "Tâche planifiée (node-cron) + vérification trips avec departureTime - 30min + envoi notification push + notification in-app", "P1", "Moyenne"),
    ("F", "Notifications par SMS (Orange/Free)", "Envoyer un SMS de confirmation de réservation et rappel pour les passagers sans connexion internet stable.", "Intégration API SMS (Orange SMS API ou Twilio) + templates SMS + file d'envoi + logs de livraison", "P3", "Élevée"),

    # Category: PAIEMENT
    ("CATEGORIE", "💳 PAIEMENTS & FACTURATION", None),
    ("F", "Paiement Wave intégré (sans redirection)", "Intégrer le paiement Wave directement dans l'app via Orange Money API au lieu de rediriger vers le lien Wave Business.", "Orange Money API / Wave Senegal API + webhook de confirmation + mode sandbox pour tests + gestion des timeouts", "P1", "Élevée"),
    ("F", "Paiement Orange Money", "Ajouter Orange Money comme méthode de paiement alternative à Wave pour les passagers qui utilisent Orange.", "Orange Money Developer API + même flow que Wave + configuration du numéro marchand + gestion de la validation", "P2", "Élevée"),
    ("F", "Portefeuille virtuel (Wallet)", "Permettre aux utilisateurs de recharger un solde et payer directement depuis leur portefeuille BOKKO.", "Nouveau modèle Wallet (userId, balance, transactions) + API /api/wallet/* + interface de recharge + historique transactions", "P3", "Élevée"),
    ("F", "Historique des paiements", "Vue dédiée listant tous les paiements effectués (réservations + colis) avec filtres par date, statut, méthode.", "API /api/payments/history + composant PaymentHistory + filtres + export PDF optionnel + totaux par période", "P2", "Moyenne"),
    ("F", "Reçu de paiement PDF", "Générer un reçu PDF téléchargeable pour chaque paiement confirmé (réservation ou colis).", "Bibliothèque jspdf ou puppeteer + template de reçu BOKKO + logo + détails trajet/colis + montant + méthode", "P3", "Moyenne"),
    ("F", "Frais de service / Commission", "Appliquer une commission BOKKO sur chaque transaction (ex: 5%) avec tableau de bord des revenus.", "Modèle Transaction + calcul automatique commission + dashboard admin des revenus + config du pourcentage", "P3", "Moyenne"),

    # Category: EXPÉRIENCE PASSAGER
    ("CATEGORIE", "🧑‍💼 EXPÉRIENCE PASSAGER", None),
    ("F", "Filtres avancés de recherche", "Filtrer les trajets par prix (min-max), heure de départ, nombre de places, note du chauffeur, type de trajet.", "Ajout paramètres dans /api/trips GET + composant de filtres expandable + UI sliders/pickers + sauvegarde des filtres", "P1", "Moyenne"),
    ("F", "Favoris (chauffeurs & trajets)", "Sauvegarder ses chauffeurs préférés et recevoir une notification quand ils publient un nouveau trajet.", "Nouveaux modèles FavoriteDriver, FavoriteRoute + API CRUD + composant FavorisView + icône étoile sur les cartes trajet", "P2", "Moyenne"),
    ("F", "Recherche de trajets réguliers", "Permettre au passager de configurer un trajet récurrent (ex: tous les lundis Thiès→Dakar) et recevoir des alertes.", "Modèle RecurringSearch (origin, destination, days[], active) + tâche planifiée quotidienne + notification quand un trajet match", "P2", "Moyenne"),
    ("F", "Historique complet des trajets", "Page regroupant tous les trajets réservés (passés, en cours, à venir) avec statistiques personnelles.", "API /api/passengers/history + composant TripHistory + filtres temporels + stats (nb trajets, km, dépenses)", "P2", "Faible"),
    ("F", "Partage d'expérience / Avis passager", "Le passager peut laisser un commentaire détaillé (pas juste une note) sur son expérience de trajet visible publiquement.", "Enrichir modèle Rating (add reviewText, addResponse from driver) + composant ReviewDetail + affichage sur profil chauffeur", "P3", "Faible"),
    ("F", "Chat intégré passager-chauffeur", "Messagerie instantanée entre passager et chauffeur pour coordonner le point de rendez-vous sans utiliser WhatsApp.", "WebSocket (Socket.io) ou polling + modèle Message + composant ChatView + historique + notifications nouveaux messages", "P2", "Élevée"),
    ("F", "Estimation prix dynamique", "Afficher une estimation du prix en temps réel basée sur la distance, la demande et les tarifs du marché.", "Calcul distance OSRM + algorithme de pricing dynamique + API /api/trips/estimate + affichage dans le formulaire de recherche", "P3", "Moyenne"),

    # Category: EXPÉRIENCE CHAUFFEUR
    ("CATEGORIE", "🚕 EXPÉRIENCE CHAUFFEUR", None),
    ("F", "Réservation en attente : auto-refus", "Refuser automatiquement les réservations en attente si le chauffeur ne répond pas dans les 2 heures.", "Tâche planifiée (node-cron) + vérification reservations PENDING + mise à jour statut → CANCELLED + notification au passager", "P1", "Moyenne"),
    ("F", "Annulation par le chauffeur", "Le chauffeur peut annuler une réservation confirmée avec motif obligatoire (pas de place, problème mécanique, etc.).", "Modification API /api/reservations/[id] PATCH + motif d'annulation + notification passager + remboursement automatique si payé", "P1", "Faible"),
    ("F", "Tableau de bord avancé", "Statistiques détaillées : revenus par jour/semaine/mois, nombre de trajets, taux de remplissage, notes moyennes, graphiques.", "API /api/stats/detailed + composant DashboardStats avec graphiques (Chart.js/Recharts) + filtres temporels + export CSV", "P2", "Moyenne"),
    ("F", "Répétition de trajet", "Le chauffeur peut dupliquer un trajet précédent pour le republier rapidement avec de nouvelles dates/heures.", "API /api/trips/[id]/duplicate + pré-remplir le formulaire DriverPublish + modification date/heure/price si souhaité", "P2", "Faible"),
    ("F", "Gestion des allers-retours", "Créer un trajet aller-retour en une seule opération (ex: Thiès→Dakar le matin, Dakar→Thiès le soir).", "Enrichir le formulaire DriverPublish + option aller-retour + création de 2 trips liés + affichage groupé dans DriverTrips", "P2", "Moyenne"),
    ("F", "Photo du véhicule upload", "Permettre au chauffeur d'ajouter des photos de son véhicule (intérieur et extérieur) visibles par les passagers.", "Upload vers stockage cloud (Cloudinary/S3) ou base64 optimisé + modèle VehiclePhoto + galerie photos + compression auto", "P3", "Moyenne"),
    ("F", "Calendrier de disponibilité", "Le chauffeur définit ses jours et heures de disponibilité pour recevoir des alertes de demande adaptées.", "Modèle Availability (dayOfWeek, startTime, endTime) + composant CalendarEditor + filtrage dans la recherche passager", "P3", "Moyenne"),
    ("F", "Réponse aux avis", "Le chauffeur peut répondre publiquement aux avis/notes laissés par les passagers sur son profil.", "Modèle RatingReply + API /api/ratings/[id]/reply + interface dans DriverProfile + modération", "P4", "Faible"),

    # Category: COLIS
    ("CATEGORIE", "📦 COLIS & LIVRAISON", None),
    ("F", "Suivi colis en temps réel", "Suivi détaillé de la livraison : Récupéré → En cours de livraison → Arrivé à la gare → Prêt à être retiré → Livré.", "Enrichir les statuts Package + maj GPS par le chauffeur + notifications à chaque étape + timeline visuel enrichi", "P1", "Moyenne"),
    ("F", "Preuve de livraison (photo)", "Le chauffeur prend une photo à la livraison comme preuve, visible par l'expéditeur et le destinataire.", "Upload photo + modèle DeliveryProof + API /api/packages/[id]/proof + affichage photo + timestamp GPS", "P2", "Moyenne"),
    ("F", "Confirmation de réception par destinataire", "Le destinataire du colis confirme la réception via un code SMS ou un lien de suivi unique.", "Génération code de suivi 6 chiffres + SMS au destinataire + API /api/packages/[id]/confirm + validation du code", "P2", "Moyenne"),
    ("F", "Prix dynamique par taille/poids", "Calcul automatique du prix du colis basé sur la taille, le poids et la distance, avec grille tarifaire configurable.", "Grille tarifaire dans le modèle Trip + calcul côté frontend et backend + affichage estimation temps réel", "P2", "Faible"),
    ("F", "Historique des colis détaillé", "Page récapitulative de tous les colis expédiés avec filtres (statut, date, destination) et statistiques.", "API /api/packages/history + composant PackageHistory + stats (nb colis, délais moyens, coûts) + export", "P3", "Faible"),

    # Category: ADMINISTRATION
    ("CATEGORIE", "⚙️ ADMINISTRATION & MODÉRATION", None),
    ("F", "Panneau administrateur", "Interface d'administration pour gérer les utilisateurs, les signalements, les statistiques globales et la modération.", "Nouveau rôle ADMIN + DashboardAdmin + gestion utilisateurs (block/unblock) + statistiques globales + modération avis", "P2", "Élevée"),
    ("F", "Signalement d'utilisateur", "Permettre aux utilisateurs de signaler un comportement inapproprié (chauffeur ou passager) avec motif et description.", "Modèle Report (reporterId, reportedId, reason, description, status) + API /api/reports + composant ReportUserForm", "P2", "Moyenne"),
    ("F", "Modération des avis", "Le modérateur/admin peut masquer ou supprimer les avis inappropriés signalés par les utilisateurs.", "Modération dans panneau admin + statut HIDDEN sur Rating + notification à l'auteur si supprimé + log de modération", "P3", "Faible"),
    ("F", "Gestion des litiges", "Système de résolution des litiges (paiement non confirmé, colis endommagé, etc.) entre passager et chauffeur.", "Modèle Dispute (reservationId/packageId, reason, status, messages[]) + API /api/disputes + interface de médiation", "P3", "Moyenne"),
    ("F", "Statistiques globales de la plateforme", "Dashboard avec KPIs : nombre d'utilisateurs, trajets, réservations, revenus totaux, taux de satisfaction, croissance.", "API /api/admin/stats + graphiques Recharts/Chart.js + filtres par période + export CSV + métriques clés", "P3", "Moyenne"),

    # Category: UX & DESIGN
    ("CATEGORIE", "🎨 UX, DESIGN & ACCESSIBILITÉ", None),
    ("F", "Mode sombre (Dark Mode)", "Basculer entre le thème clair et le thème sombre selon les préférences de l'utilisateur.", "Tailwind dark mode + toggle dans les paramètres + persistance du choix dans le store + composants adaptés", "P2", "Moyenne"),
    ("F", "Onboarding / Tutoriel", "Guide interactif à la première connexion expliquant les fonctionnalités principales de l'app étape par étape.", "Composant OnboardingSlides (3-5 slides) + stockage firstLogin dans le store + skip possible + animations", "P2", "Faible"),
    ("F", "Mode hors-ligne amélioré", "Permettre de consulter ses réservations, colis et profils même sans connexion internet grâce au cache local.", "Service Worker avancé + IndexedDB pour le cache de données + stratégie stale-while-revalidate + indicateur hors-ligne", "P2", "Élevée"),
    ("F", "Animation de transition entre vues", "Ajouter des animations fluides (slide/fade) lors de la navigation entre les différents écrans de l'app.", "Framer Motion + AnimatePresence + transition de page type slide-in/slide-out + réglage vitesse/curves", "P3", "Faible"),
    ("F", "Accessibilité (a11y)", "Rendre l'app accessible aux personnes en situation de handicap : labels ARIA, navigation clavier, contraste suffisant.", "Audit a11y + aria-labels + focus management + clavier navigation + contraste AA minimum + screen reader support", "P3", "Moyenne"),
    ("F", "Multi-langue (Wolof & Anglais)", "Ajouter le Wolof et l'Anglais en plus du Français pour toucher une audience plus large au Sénégal.", "next-intl ou i18next + fichiers de traduction FR/EN/WO + sélecteur de langue + persistance du choix", "P4", "Élevée"),

    # Category: FIDÉLITÉ & CROISSANCE
    ("CATEGORIE", "🏆 FIDÉLITÉ & CROISSANCE", None),
    ("F", "Programme de fidélité", "Système de points : 1 point par trajet/réservation, avantages pour les niveaux Bronze/Argent/Or/Platine.", "Modèle LoyaltyPoints + calcul automatique des points + niveaux + récompenses (réduction, trajet gratuit) + badge UI", "P3", "Moyenne"),
    ("F", "Code parrainage", "Chaque utilisateur a un code unique. Le parrain et le filleul reçoivent des crédits/moins lors de la 1ère réservation.", "Champ referralCode dans User + modèle Referral (referrerId, referredId, reward) + tracking + attribution automatique", "P3", "Moyenne"),
    ("F", "Promotions & codes promo", "Créer et gérer des codes promotionnels (ex: BOKKO10 pour -10%) applicables lors de la réservation.", "Modèle PromoCode (code, type, value, validFrom, validUntil, maxUses) + validation à la réservation + dashboard admin", "P3", "Moyenne"),
    ("F", "Badges & accomplissements", "Attribuer des badges visuels aux utilisateurs pour leurs actions (1er trajet, 10 trajets, 5 avis, etc.).", "Modèle Badge + modèles UserBadge + logique d'attribution automatique + galerie de badges + affichage sur profil", "P4", "Faible"),

    # Category: TECHNIQUE
    ("CATEGORIE", "🔧 TECHNIQUE & INFRASTRUCTURE", None),
    ("F", "Upload photos sur cloud (S3/Cloudinary)", "Remplacer le stockage base64 des photos par un stockage cloud avec CDN pour de meilleures performances.", "Intégration Cloudinary (gratuit 25GB) ou AWS S3 + upload sécurisé + URLs publiques + optimisation auto images", "P2", "Moyenne"),
    ("F", "Base de données PostgreSQL", "Migrer de SQLite vers PostgreSQL pour une meilleure scalabilité, concurrence et performances en production.", "Migration Prisma + déploiement PostgreSQL (Supabase/Neon) + scripts de migration de données + configuration connection pool", "P2", "Moyenne"),
    ("F", "Logging & Monitoring", "Ajouter un système de logging centralisé pour tracer les erreurs, les performances API et le comportement des utilisateurs.", "Intégration Sentry (gratuit pour projets open-source) + logger structuré + dashboard erreurs + alertes email/Slack", "P2", "Moyenne"),
    ("F", "Tests automatisés", "Écrire des tests unitaires et d'intégration pour les API critiques (auth, réservation, paiement) et les composants clés.", "Jest + React Testing Library + tests API routes + tests composants + CI pipeline (GitHub Actions) + coverage", "P2", "Élevée"),
    ("F", "Redirection de base de données (Supabase/Neon)", "Héberger la base de données en cloud pour la production avec backup automatique et haute disponibilité.", "Déploiement Supabase ou Neon (gratuit tier) + migration données + variables d'environnement + connection pooling", "P2", "Faible"),
]

# Write data
row = 10
num = 0
for feat in features:
    if feat[0] == "CATEGORIE":
        # Category row
        ws.merge_cells(f'B{row}:G{row}')
        cell = ws.cell(row=row, column=2, value=feat[1])
        cell.font = category_font
        cell.fill = category_fill
        cell.alignment = left_align
        cell.border = thin_border
        for c in range(3, 8):
            ws.cell(row=row, column=c).fill = category_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 26
    else:
        num += 1
        is_alt = num % 2 == 0
        bg = grey_alt if is_alt else white
        
        values = [num, feat[1], feat[2], feat[3], feat[4], feat[5]]
        for i, val in enumerate(values):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            
            if i == 0:  # #
                cell.font = normal_font
                cell.alignment = center_align
            elif i == 1:  # Feature
                cell.font = Font(name='Times New Roman', size=11, bold=True, color="1A1A1A")
                cell.alignment = left_top
            elif i == 2:  # Description
                cell.font = desc_font
                cell.alignment = left_top
            elif i == 3:  # Details
                cell.font = Font(name='Times New Roman', size=9, color="666666")
                cell.alignment = left_top
            elif i == 4:  # Priority
                p = val
                cell.alignment = center_align
                if p == "P1": cell.font = p1_font
                elif p == "P2": cell.font = p2_font
                elif p == "P3": cell.font = p3_font
                else: cell.font = p4_font
            elif i == 5:  # Complexity
                cell.font = normal_font
                cell.alignment = center_align
        
        # Adjust row height for content
        ws.row_dimensions[row].height = 52
    row += 1

# Summary section
row += 2
ws.merge_cells(f'B{row}:G{row}')
ws[f'B{row}'] = "RÉSUMÉ PAR PRIORITÉ"
ws[f'B{row}'].font = Font(name='Times New Roman', size=14, bold=True, color=green_primary)
row += 1

summary_headers = ["Priorité", "Nombre", "Description", "Recommandation"]
for i, h in enumerate(summary_headers):
    col = i + 2
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.merge_cells(f'E{row}:G{row}')

row += 1
summaries = [
    ("P1 - Critique", "8", "Fonctionnalités essentielles pour le lancement et la sécurité des utilisateurs", "À implémenter en priorité avant mise en production"),
    ("P2 - Important", "16", "Fonctionnalités importantes pour l'amélioration de l'UX et la rétention", "À implémenter dans les 2-3 premiers mois après le lancement"),
    ("P3 - Moyen", "13", "Fonctionnalités d'enrichissement et de différenciation", "À planifier sur le roadmap trimestrielle"),
    ("P4 - Futur", "3", "Fonctionnalités avancées pour la croissance à long terme", "À étudier quand la base d'utilisateurs sera suffisante"),
]

for idx, s in enumerate(summaries):
    bg = grey_alt if idx % 2 == 0 else white
    ws.cell(row=row, column=2, value=s[0]).font = Font(name='Times New Roman', size=11, bold=True)
    ws.cell(row=row, column=3, value=s[1]).font = normal_font
    ws.cell(row=row, column=3).alignment = center_align
    ws.merge_cells(f'D{row}:E{row}')
    ws.cell(row=row, column=4, value=s[2]).font = normal_font
    ws.merge_cells(f'F{row}:G{row}')
    ws.cell(row=row, column=6, value=s[3]).font = desc_font
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).alignment = left_align
    ws.cell(row=row, column=3).alignment = center_align
    row += 1

# Save
wb.save('/home/z/my-project/download/BOKKO_Fonctionnalites_A_Ajouter.xlsx')
print("OK")
